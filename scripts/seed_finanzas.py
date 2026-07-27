"""Seed de plan de cuentas, personal, gastos fijos y supuestos financieros
leyendo directamente el Archivo Maestro de Excel (hojas 04_Plan_Cuentas,
06_Personal, 07_Gastos_Fijos). Requiere que scripts/seed_catalogo.py ya se
haya corrido antes (necesita categorías y familias existentes).

Uso:
    python scripts/seed_finanzas.py ["ruta/al/Archivo Maestro.xlsx"]

Es idempotente: si ya fue sembrado (marca en meta), no hace nada.
"""
from __future__ import annotations

import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

import openpyxl

from aglegal import db
from aglegal.db import now_iso
from aglegal.repositories import Repository

SEED_KEY = "finanzas_seed_v1"
DEFAULT_XLSX = ROOT / "AG Legal Archivo Maestro2026.xlsx"
YEAR = "2026"  # el archivo maestro solo cubre julio-diciembre 2026


def _cell(ws, row: int, col: str) -> str:
    v = ws[f"{col}{row}"].value
    return str(v).strip() if v is not None else ""


def _mes(value: str) -> str | None:
    v = (value or "").strip()
    if not v:
        return None
    return f"{YEAR}-{int(float(v)):02d}"


def _seed_plan_cuentas(repo: Repository, wb, category_ids: dict[str, int], family_codes_to_id: dict[str, int]) -> int:
    ws = wb["04_Plan_Cuentas"]
    count = 0
    row = 5
    while _cell(ws, row, "A"):
        account_code = _cell(ws, row, "A")
        tipo_es = _cell(ws, row, "B")
        grupo = _cell(ws, row, "C")
        subgrupo = _cell(ws, row, "D")
        nombre = _cell(ws, row, "E")
        naturaleza = _cell(ws, row, "F")
        cat_code = _cell(ws, row, "G")
        fam_code = _cell(ws, row, "H")
        centro_costo = _cell(ws, row, "I")
        afecta = _cell(ws, row, "J")
        regla = _cell(ws, row, "L")

        repo.create_cuenta(
            account_code=account_code,
            tipo=tipo_es,
            grupo=grupo,
            subgrupo=subgrupo,
            nombre=nombre,
            naturaleza=naturaleza,
            category_id=category_ids.get(cat_code),
            family_id=family_codes_to_id.get(fam_code),
            centro_costo=centro_costo,
            afecta_utilidad=(afecta != "No"),
            regla_de_uso=regla,
            created_at=now_iso(),
        )
        count += 1
        row += 1
    return count


def _seed_personal(repo: Repository, wb) -> int:
    ws = wb["06_Personal"]
    count = 0
    row = 5
    while _cell(ws, row, "A"):
        persona = _cell(ws, row, "B")
        cargo = _cell(ws, row, "C")
        monto = _cell(ws, row, "D")
        mes_inicio = _mes(_cell(ws, row, "E"))
        mes_fin = _mes(_cell(ws, row, "F"))
        repo.create_persona(
            persona=persona, cargo=cargo, monto_mensual_text=monto,
            mes_inicio=mes_inicio, mes_fin=mes_fin, created_at=now_iso(),
        )
        count += 1
        row += 1
    return count


def _seed_gastos_fijos(repo: Repository, wb) -> int:
    ws = wb["07_Gastos_Fijos"]
    count = 0
    row = 5
    while _cell(ws, row, "A"):
        concepto = _cell(ws, row, "B")
        tipo = _cell(ws, row, "C")
        monto = _cell(ws, row, "D")
        mes_inicio = _mes(_cell(ws, row, "E"))
        mes_fin = _mes(_cell(ws, row, "F"))
        repo.create_gasto_fijo(
            concepto=concepto, tipo=tipo, monto_mensual_text=monto,
            mes_inicio=mes_inicio, mes_fin=mes_fin, created_at=now_iso(),
        )
        count += 1
        row += 1
    return count


def _seed_supuestos(repo: Repository, wb) -> None:
    ws = wb["07_Gastos_Fijos"]
    costo_variable = float(ws["B12"].value)
    margen_operativo_meta = float(ws["B13"].value)
    margen_seguridad = float(ws["B14"].value)
    repo.create_supuestos(
        periodo=YEAR,
        costo_variable_pct=costo_variable,
        margen_operativo_meta_pct=margen_operativo_meta,
        margen_seguridad_pct=margen_seguridad,
        created_at=now_iso(),
    )


def main() -> int:
    xlsx_path = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_XLSX
    if not xlsx_path.exists():
        print(f"No se encontró el archivo: {xlsx_path}")
        return 1

    conn = db.connect()
    try:
        db.init_db(conn)
        existing = conn.execute("SELECT value FROM meta WHERE key=%s", (SEED_KEY,)).fetchone()
        if existing:
            print("El plan de cuentas / personal / gastos fijos ya fue sembrado — no se hace nada.")
            return 0

        catalogo_seeded = conn.execute("SELECT value FROM meta WHERE key='catalogo_seed_v1'").fetchone()
        if not catalogo_seeded:
            print("Corre primero scripts/seed_catalogo.py — este script necesita categorías y familias existentes.")
            return 1

        repo = Repository(conn)
        category_ids = {r["category_code"]: r["id"] for r in repo.list_categorias()}
        family_codes_to_id = {r["family_code"]: r["id"] for r in repo.list_familias()}

        wb = openpyxl.load_workbook(str(xlsx_path), data_only=True)

        cuentas_count = _seed_plan_cuentas(repo, wb, category_ids, family_codes_to_id)
        print(f"Plan de cuentas: {cuentas_count}")

        personal_count = _seed_personal(repo, wb)
        print(f"Personal: {personal_count}")

        gastos_count = _seed_gastos_fijos(repo, wb)
        print(f"Gastos fijos: {gastos_count}")

        _seed_supuestos(repo, wb)
        print("Supuestos financieros: 1 (periodo 2026)")

        conn.execute(
            "INSERT INTO meta(key, value) VALUES(%s,%s) ON CONFLICT(key) DO UPDATE SET value=excluded.value",
            (SEED_KEY, now_iso()),
        )
        conn.commit()
        print("Plan de cuentas, personal y gastos fijos sembrados desde el Archivo Maestro.")
        return 0
    finally:
        conn.close()


if __name__ == "__main__":
    raise SystemExit(main())
